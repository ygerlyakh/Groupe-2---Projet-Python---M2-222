#%% IMPORTS
import pandas as pd
import numpy as np
import yfinance as yf
import matplotlib.pyplot as plt
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Progress bar (tqdm) with fallback
try:
    from tqdm.auto import tqdm
except Exception:
    tqdm = None


#%% PARAMETERS
MASTER_CSV_PATH = "S&P 500 Historical Components & Changes(11-16-2025).csv"
GLOBAL_START_DOWNLOAD = "2014-01-01"
GLOBAL_END_DOWNLOAD   = "2025-11-16"

MOMENTUM_WINDOW = 126
MOMENTUM_GAP    = 20
TOP_K           = 20
LEVERAGE        = 1.0
VOL_WINDOW      = 20


#%% HELPERS
def _downloads_dir() -> Path:
    d = Path.home() / "Downloads"
    return d if d.exists() else Path.cwd()

def _pbar(it, desc: str):
    if tqdm is None:
        print(desc)
        return it
    return tqdm(it, desc=desc)


#%% 1. DATA DOWNLOAD (dynamic universe)
def get_dynamic_universe(csv_path, end_date):
    try:
        print("1) Reading universe CSV...")
        df_master = pd.read_csv(csv_path, parse_dates=["date"], index_col="date")
        df_master["tickers"] = df_master["tickers"].astype(str).apply(lambda x: [t.strip() for t in x.split(',') if t.strip()])
        df_master = df_master.sort_index().loc[GLOBAL_START_DOWNLOAD:end_date]

        all_tickers = set(t for sublist in df_master["tickers"] for t in sublist)
        tickers_yf = [x.replace('.', '-') for x in sorted(all_tickers)]

        print(f"2) Downloading prices for {len(tickers_yf)} tickers from yfinance...")
        data = yf.download(
            tickers=tickers_yf,
            start=GLOBAL_START_DOWNLOAD,
            end=end_date,
            progress=True,           # yfinance built-in progress
            auto_adjust=False,
            threads=True
        )

        # Robust price extraction
        if isinstance(data.columns, pd.MultiIndex):
            fields = data.columns.get_level_values(0).unique().tolist()
            if "Adj Close" in fields:
                prices = data["Adj Close"]
            elif "Close" in fields:
                prices = data["Close"]
            else:
                raise KeyError(f"No price field found. Fields: {fields}")
        else:
            prices = data["Adj Close"] if "Adj Close" in data.columns else data

        prices = prices.dropna(axis=1, how="all")
        if prices.empty:
            raise RuntimeError("Downloaded prices are empty after cleaning.")

        print("3) Building dynamic membership mask...")
        mask = pd.DataFrame(False, index=prices.index, columns=prices.columns)
        common_dates = prices.index.intersection(df_master.index)

        for date in _pbar(common_dates, desc="Masking membership dates"):
            valid_tickers = [t.replace('.', '-') for t in df_master.loc[date, "tickers"]]
            cols = list(set(valid_tickers) & set(prices.columns))
            if cols:
                mask.loc[date, cols] = True

        # forward-fill membership and prices
        mask = mask.replace(False, np.nan).ffill().fillna(False).astype(bool)
        return prices.where(mask).ffill()

    except Exception as e:
        print(f"Error in get_dynamic_universe: {e}")
        return pd.DataFrame()


#%% 2. SIGNAL CALCULATION (VOL-ADAPT)
def get_current_allocation(prices):
    if prices.empty:
        return pd.DataFrame(), None

    mom_series = (prices.shift(MOMENTUM_GAP) / prices.shift(MOMENTUM_WINDOW + MOMENTUM_GAP)) - 1

    # Annualized realized vol (as in your code)
    vol_series = prices.pct_change().rolling(VOL_WINDOW).std() * np.sqrt(252)

    # Use last date in prices (as in your code)
    last_dt = prices.index[-1]

    last_mom = mom_series.loc[last_dt].dropna()
    last_vol = vol_series.loc[last_dt].reindex(last_mom.index)

    # Fill missing/zero vol safely using median positive vol
    median_vol = last_vol[last_vol > 0].median()
    if pd.isna(median_vol) or median_vol <= 0:
        # Fallback if median can't be computed
        median_vol = last_vol.replace(0, np.nan).dropna().median()
    if pd.isna(median_vol) or median_vol <= 0:
        # absolute fallback (should be rare)
        median_vol = 0.20  # 20% annualized as a last resort

    last_vol = last_vol.fillna(median_vol).replace(0, median_vol)

    print(f"\nAnalysis Date: {last_dt.date()}")
    if len(last_mom) < TOP_K * 2:
        return pd.DataFrame(), None

    sorted_scores = last_mom.sort_values(ascending=False)
    longs  = sorted_scores.head(TOP_K)
    shorts = sorted_scores.tail(TOP_K).sort_values(ascending=True)

    # VOL-ADAPT weights: inverse-vol within each leg, scaled to LEVERAGE
    v_long = last_vol.loc[longs.index]
    inv_v_long = 1.0 / v_long
    w_long = (inv_v_long / inv_v_long.sum()) * LEVERAGE

    v_short = last_vol.loc[shorts.index]
    inv_v_short = 1.0 / v_short
    w_short = (inv_v_short / inv_v_short.sum()) * (-LEVERAGE)

    df_long = pd.DataFrame({
        "Rank": range(1, TOP_K + 1),
        "Ticker": longs.index,
        "Side": "LONG",
        "Historical Return": longs.values,
        "Weight": w_long.values,
        "Vol": v_long.values
    })

    df_short = pd.DataFrame({
        "Rank": range(1, TOP_K + 1),
        "Ticker": shorts.index,
        "Side": "SHORT",
        "Historical Return": shorts.values,
        "Weight": w_short.values,
        "Vol": v_short.values
    })

    return pd.concat([df_long, df_short], ignore_index=True), last_dt


#%% 3) EXPORT EXCEL (save ONLY to Downloads)
def export_excel_dashboard(df_alloc, date_obj):
    if df_alloc.empty:
        return None

    cols = ["Rank", "Ticker", "Historical Return", "Weight", "Vol"]
    df_long = df_alloc[df_alloc["Side"] == "LONG"][cols].reset_index(drop=True)
    df_short = df_alloc[df_alloc["Side"] == "SHORT"][cols].reset_index(drop=True)

    filename = _downloads_dir() / f"Allocation_{date_obj.date()}.xlsx"

    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        wb = writer.book
        ws = wb.create_sheet("Allocation")
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        align_center = Alignment(horizontal="center", vertical="center")
        thin_border = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
        font_white = Font(color="FFFFFF", bold=True)
        font_bold = Font(bold=True)
        fill_grey = PatternFill(start_color="f2f2f2", fill_type="solid")

        ws.merge_cells("A1:K1")
        ws["A1"].value = f"VOL-ADAPTED PORTFOLIO - {date_obj.date()}"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = align_center

        def create_table(dataframe, start_col, title, color_hex):
            ws.cell(row=3, column=start_col, value=title).font = Font(bold=True, color=color_hex)
            fill_header = PatternFill(start_color=color_hex, fill_type="solid")

            # header
            for i, col_name in enumerate(dataframe.columns):
                c = ws.cell(row=4, column=start_col + i, value=col_name)
                c.fill = fill_header
                c.font = font_white
                c.border = thin_border
                c.alignment = align_center

            # rows
            for r, row in dataframe.iterrows():
                values = [row["Rank"], row["Ticker"], row["Historical Return"], row["Weight"], row["Vol"]]
                for i, val in enumerate(values):
                    c = ws.cell(row=5 + r, column=start_col + i, value=float(val) if i >= 2 else val)
                    c.border = thin_border
                    c.alignment = align_center
                    if i >= 2:
                        c.number_format = "0.00%"

            # totals row
            row_total = 5 + len(dataframe)
            for i in range(5):
                c = ws.cell(row=row_total, column=start_col + i)
                c.fill = fill_grey
                c.border = thin_border
                c.alignment = align_center

                if i == 1:
                    c.value = "TOTAL"
                    c.font = font_bold
                if i == 3:
                    c.value = float(dataframe["Weight"].sum())
                    c.number_format = "0.00%"
                    c.font = font_bold

        create_table(df_long, start_col=1, title="BUY / LONG", color_hex="27ae60")
        create_table(df_short, start_col=7, title="SELL / SHORT", color_hex="c0392b")

    print(f"Excel saved to: {filename}")
    return filename


#%% 4) DASHBOARD IMAGE (DISPLAY ONLY — DO NOT SAVE)
def export_dashboard_display_only(df_alloc, date_obj):
    """
    Generates the dashboard image and shows it, but DOES NOT save a PNG.
    """
    if df_alloc.empty:
        return

    df = df_alloc.copy()

    # Format display values (strings)
    for col in ["Historical Return", "Weight"]:
        df[col] = df[col].apply(lambda x: f"{x:.2%}")
    df["Vol"] = df["Vol"].apply(lambda x: f"{x:.1%}")

    cols = ["Rank", "Ticker", "Historical Return", "Weight", "Vol"]

    fig, axes = plt.subplots(1, 2, figsize=(14, 12), dpi=150)
    fig.patch.set_facecolor("white")

    plt.suptitle(f"VOL-ADAPTED PORTFOLIO - {date_obj.date()}", fontsize=20, fontweight="bold", y=0.98)

    configs = [
        (axes[0], df[df["Side"] == "LONG"][cols], "BUY / LONG", "#27ae60"),
        (axes[1], df[df["Side"] == "SHORT"][cols], "SELL / SHORT", "#c0392b"),
    ]

    for ax, data, title, color in configs:
        ax.axis("off")
        ax.set_title(title, fontsize=14, fontweight="bold", color=color)

        if data.empty:
            continue

        table = ax.table(cellText=data.values, colLabels=data.columns, loc="center", cellLoc="center")
        table.auto_set_font_size(False)
        table.set_fontsize(10)
        table.scale(1, 1.8)

        for (row, col), cell in table.get_celld().items():
            cell.set_edgecolor("white")
            if row == 0:
                cell.set_facecolor(color)
                cell.set_text_props(weight="bold", color="white")
            else:
                cell.set_facecolor("#f4f4f4")

    plt.tight_layout()
    plt.show()


#%% 5. MAIN EXECUTION
if __name__ == "__main__":
    prices = get_dynamic_universe(MASTER_CSV_PATH, GLOBAL_END_DOWNLOAD)
    df_alloc, last_date = get_current_allocation(prices)

    if not df_alloc.empty:
        # Progress print for export stage
        print("\nExporting Excel to Downloads...")
        export_excel_dashboard(df_alloc, last_date)

        print("Generating dashboard image (display only)...")
        export_dashboard_display_only(df_alloc, last_date)
    else:
        print("No allocation produced (df_alloc is empty).")
